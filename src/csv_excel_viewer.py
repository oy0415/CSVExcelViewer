#csvファイルとExcelを結合するプログラム
import csv
import TkEasyGUI as sg
# ★追加：Excel対応
import openpyxl as xl
import datetime
import os


def main():
    while True:
        #CSVファイルを選ぶ
        files = sg.popup_get_file(
            "CSVファイルを複数選択",
            multiple_files=True, #複数ファイルの選択
            no_window=True,
            file_types=(
                ("CSV/Excel", "*.csv;*.xlsx;*.xlsm"),
                ("CSVファイル", "*.csv"),
                ("Excelファイル", "*.xlsx;*.xlsm"),
                )
        )
        #追加1：複数選択も戻り値をlistへ正規化
        if not files:
            break
        if isinstance(files, str):
            files = [p for p in files.split(";") if p.strip()]
        if len(files) == 0:
            break
        
        #複数のファイルをまとめる
        all_data = []
        for filename in files:
            data = read_table(filename)
            if data is None:
                sg.popup_error(filename + "読み込めません")
                continue
            #もしヘッダーが同じなら省略する
            if len(all_data) >= 2 and len(data) >= 2:
                if all_data[0] == data[0]:
                    data = data[1:]
            all_data += data
        #結合したデータをテーブルに表示する
        if show_csv(all_data) == False:
            break

# ★追加：拡張子に応じてリーダーを切り替え
def read_table(filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        return read_csv(filename)
    if ext in (".xlsx", ".xlsm"):
        return read_excel(filename) 
    return None

# ★追加：Excel(.xlsx/.xlsm)を読む
def read_excel(filename, sheet_name=None):
    wb = xl.load_workbook(filename, data_only=True, read_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        cleaned = []
        for c in row:
            if c is None:
                cleaned.append("")
            elif isinstance(c, (datetime.datetime, datetime.date)):
                # 時刻が0:00でなければ日時、それ以外は日付
                if isinstance(c, datetime.datetime) and (c.time() != datetime.time(0, 0)):
                    cleaned.append(c.strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    cleaned.append(c.strftime("%Y-%m-%d"))
            else:
                cleaned.append(str(c).strip())
        # 全セル空はスキップ
        if not any(cleaned):
            continue
        data.append(cleaned)
    return data
        
#CSVファイルを読む - UTF-8/Shift_JIS(CP932)対応版
#変更2:read_csv を強化：区切り自動判定・BOM/空白除去・空行スキップ
def read_csv(filename):
    import csv
    encodings = ["utf-8-sig", "UTF-8", "CP932", "EUC-JP"]
    for enc in encodings:
        try:
            with open(filename, "r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                data = []
                for row in reader:
                    cleaned = [str(c).strip().lstrip("\ufeff") for c in row]
                    # ★ 空行はスキップ（全セルが空/空白）
                    if not any(cleaned):
                        continue
                    data.append(cleaned)
                return data
        except Exception:
            continue
    return None

#CSVをテーブルに表示する
def show_csv(data):
    # ★追加：行ごとの列数を最大列数に合わせる（Table要件）
    max_cols = max(len(r) for r in data)
    data = [r + [""] * (max_cols - len(r)) for r in data]

    if len(data) == 0:
        data = [["空"], ["空"]]
    #レイアウトを定義
    layout = [
        [sg.Table(
            key='-table-',
            values=data[1:],
            headings=data[0],
            expand_x=True, expand_y=True,
            justification='left',
            auto_size_columns=True,
            max_col_width=30,
            font=("Arial", 14))],
        [sg.Button('ファイル選択'), sg.Button('保存'), sg.Button('終了')]
        ]
    #ウィンドウを作成
    window = sg.Window("CSVビュワー", layout,
                       size=(500, 300), resizable=True, finalize=True)
    #イベントのループ
    flag_continue = False
    while True:
        event, _ = window.read()
        if event in [sg.WINDOW_CLOSED, "終了"]:
            break
        #ファイル追加ボタンを押したとき
        if event == "ファイル選択":
            flag_continue = True
            break
        #保存ボタンを押したとき
        if event == "保存":
            #ファイルを選ぶ
            filename = sg.popup_get_file(
                "保存先のCSVファイルを指定",
                save_as=True,
                no_window=True,
                file_types=(("CSVファイル", "*.csv"),)
            )
            if filename == "" or filename is None:
                continue
            #CSVファイルに書き込む
            with open(filename, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f, lineterminator="\n")
                writer.writerows(data)
    window.close()
    return flag_continue

if __name__ == "__main__":
    main()
            